import os
import json
import io
import logging
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass
import time
import pickle
import uuid
import threading

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import PyPDF2
from docx import Document
import openpyxl
from langchain_groq import ChatGroq
from flask import Flask, request

# LOGGING
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('drive_organizer.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# CONFIGURATION
SCOPES = ['https://www.googleapis.com/auth/drive']

CATEGORIES = [
    "HR",
    "Finance",
    "Academics",
    "Projects",
    "Marketing",
    "Personal",
    "Miscellaneous"
]

CONFIDENCE_THRESHOLD = 0.7
MAX_CONTENT_LENGTH = 3000

# MIME TYPES TO SKIP
SKIP_MIME_PREFIXES = (
    "image/",
    "video/",
    "audio/"
)

# Flask app for webhooks
app = Flask(__name__)
organizer_instance = None


# DATA MODELS
@dataclass
class FileInfo:
    id: str
    name: str
    mime_type: str
    size: int
    created_time: str
    parents: List[str]


@dataclass
class Classification:
    category: str
    confidence: float
    reasoning: str
    subcategory: Optional[str] = None


# GOOGLE DRIVE CLIENT
class GoogleDriveClient:
    def __init__(self, credentials_path: str):
        self.credentials_path = credentials_path
        self.service = None
        self._authenticate()

    def _authenticate(self):
        creds = None

        if os.path.exists("token.pickle"):
            with open("token.pickle", "rb") as token:
                creds = pickle.load(token)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_path, SCOPES
                )
                creds = flow.run_local_server(port=0)

            with open("token.pickle", "wb") as token:
                pickle.dump(creds, token)

        self.service = build("drive", "v3", credentials=creds)
        logger.info("✓ Authenticated with Google Drive")

    def list_files(self, folder_id: str = "root", page_size: int = 100) -> List[FileInfo]:
        files = []
        page_token = None

        try:
            while True:
                response = self.service.files().list(
                    q=f"'{folder_id}' in parents and trashed=false",
                    pageSize=page_size,
                    fields="nextPageToken, files(id, name, mimeType, size, createdTime, parents)",
                    pageToken=page_token
                ).execute()

                for file in response.get("files", []):
                    files.append(FileInfo(
                        id=file["id"],
                        name=file["name"],
                        mime_type=file["mimeType"],
                        size=int(file.get("size", 0)),
                        created_time=file["createdTime"],
                        parents=file.get("parents", [])
                    ))

                page_token = response.get("nextPageToken")
                if not page_token:
                    break

            logger.info(f"Found {len(files)} files in folder")
            return files

        except Exception as e:
            logger.error(f"Error listing files: {str(e)}")
            return []

    def get_file(self, file_id: str) -> Optional[FileInfo]:
        """Get a single file by ID"""
        try:
            file = self.service.files().get(
                fileId=file_id,
                fields="id, name, mimeType, size, createdTime, parents"
            ).execute()
            
            return FileInfo(
                id=file["id"],
                name=file["name"],
                mime_type=file["mimeType"],
                size=int(file.get("size", 0)),
                created_time=file["createdTime"],
                parents=file.get("parents", [])
            )
        except Exception as e:
            logger.error(f"Error getting file {file_id}: {str(e)}")
            return None

    def create_folder(self, folder_name: str, parent_id: str = "root") -> Optional[str]:
        try:
            # Check if folder exists
            query = (
                f"name='{folder_name}' "
                f"and mimeType='application/vnd.google-apps.folder' "
                f"and '{parent_id}' in parents and trashed=false"
            )

            response = self.service.files().list(q=query, fields="files(id)").execute()
            folders = response.get("files", [])

            if folders:
                logger.info(f"Folder '{folder_name}' already exists")
                return folders[0]["id"]

            # Create new folder
            folder = self.service.files().create(
                body={
                    "name": folder_name,
                    "mimeType": "application/vnd.google-apps.folder",
                    "parents": [parent_id]
                },
                fields="id"
            ).execute()

            logger.info(f"✓ Created folder '{folder_name}'")
            return folder["id"]

        except Exception as e:
            logger.error(f"Error creating folder '{folder_name}': {str(e)}")
            return None

    def move_file(self, file_id: str, folder_id: str) -> bool:
        try:
            file = self.service.files().get(
                fileId=file_id, fields="parents"
            ).execute()

            previous_parents = ",".join(file.get("parents", []))

            self.service.files().update(
                fileId=file_id,
                addParents=folder_id,
                removeParents=previous_parents,
                fields="id, parents"
            ).execute()

            return True

        except Exception as e:
            logger.error(f"Error moving file {file_id}: {str(e)}")
            return False

    def download_file_content(self, file_id: str, mime_type: str) -> Optional[bytes]:
        try:
            if "google-apps" in mime_type:
                request = self.service.files().export_media(
                    fileId=file_id,
                    mimeType="text/plain"
                )
            else:
                request = self.service.files().get_media(fileId=file_id)

            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request)

            done = False
            while not done:
                _, done = downloader.next_chunk()

            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Error downloading file {file_id}: {str(e)}")
            return None


# CONTENT EXTRACTION
class ContentExtractor:

    @staticmethod
    def extract(content: bytes, mime_type: str, filename: str) -> str:
        try:
            if "pdf" in mime_type:
                return ContentExtractor._from_pdf(content)
            elif "word" in mime_type or "document" in mime_type:
                return ContentExtractor._from_docx(content)
            elif "sheet" in mime_type or "excel" in mime_type:
                return ContentExtractor._from_excel(content)
            elif "text" in mime_type:
                return content.decode("utf-8", errors="ignore")[:MAX_CONTENT_LENGTH]
            else:
                return f"Filename: {filename}"
        except Exception as e:
            logger.error(f"Error extracting content: {str(e)}")
            return f"Filename: {filename}"

    @staticmethod
    def _from_pdf(content: bytes) -> str:
        reader = PyPDF2.PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages[:5]:
            text += page.extract_text() or ""
        return text[:MAX_CONTENT_LENGTH]

    @staticmethod
    def _from_docx(content: bytes) -> str:
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs)[:MAX_CONTENT_LENGTH]

    @staticmethod
    def _from_excel(content: bytes) -> str:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        text = "Sheets: " + ", ".join(wb.sheetnames) + "\n"
        text += " ".join(
            str(cell.value) for row in list(sheet.iter_rows(max_row=20))
            for cell in row if cell.value
        )
        return text[:MAX_CONTENT_LENGTH]


# AI CLASSIFIER
class AIClassifier:

    def __init__(self, api_key: str):
        self.llm = ChatGroq(
            model="llama-3.3-70b-versatile",
            groq_api_key=api_key,
            temperature=0
        )

    def classify(self, file: FileInfo, content: str) -> Classification:
        try:
            prompt = self._prompt(file, content)
            response = self.llm.invoke(prompt).content.strip()
            
            # Clean up response
            response = response.strip("```json").strip("```").strip()
            result = json.loads(response)

            return Classification(
                category=result["category"],
                confidence=float(result["confidence"]),
                reasoning=result["reasoning"],
                subcategory=result.get("subcategory")
            )

        except Exception as e:
            logger.error(f"Error classifying file: {str(e)}")
            return Classification(
                category="Miscellaneous",
                confidence=0.5,
                reasoning=f"Classification error: {str(e)}"
            )

    def _prompt(self, file: FileInfo, content: str) -> str:
        return f"""
Classify this file into ONE category from:
{', '.join(CATEGORIES)}

File name: {file.name}
Type: {file.mime_type}
Size: {file.size} bytes

Content preview:
{content[:MAX_CONTENT_LENGTH]}

Return ONLY valid JSON (no markdown, no explanations):
{{
  "category": "one of the categories above",
  "confidence": 0.0-1.0,
  "reasoning": "brief explanation",
  "subcategory": "optional subcategory"
}}
"""


# WEBHOOK HANDLERS
@app.route('/webhook/drive', methods=['POST'])
def drive_webhook():
    """Handle Google Drive push notifications"""
    
    try:
        # Get notification headers
        resource_state = request.headers.get('X-Goog-Resource-State')
        resource_id = request.headers.get('X-Goog-Resource-ID')
        channel_id = request.headers.get('X-Goog-Channel-ID')
        
        logger.info(f"Webhook received: state={resource_state}, resource={resource_id}, channel={channel_id}")
        
        # Only process 'change' or 'update' notifications
        if resource_state in ['change', 'update']:
            # Process in background thread to respond quickly
            threading.Thread(target=process_new_files, daemon=True).start()
        
        return 'OK', 200
        
    except Exception as e:
        logger.error(f"Error handling webhook: {str(e)}")
        return 'Error', 500


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return {
        'status': 'running',
        'timestamp': datetime.now().isoformat(),
        'organizer': 'active' if organizer_instance else 'inactive'
    }, 200


def process_new_files():
    """Process any new unorganized files"""
    if not organizer_instance:
        logger.error("Organizer not initialized")
        return
        
    try:
        logger.info("Checking for new files to organize...")
        
        # Get all files in root
        files = organizer_instance.drive.list_files("root")
        organized_count = 0
        
        for file in files:
            # Skip if file type should be skipped
            if organizer_instance._should_skip_file(file):
                continue
            
            # Skip if already in a category folder
            if organizer_instance._is_organized(file):
                continue
            
            # Organize this file
            try:
                organized = organizer_instance.organize_single_file(file)
                if organized:
                    organized_count += 1
            except Exception as e:
                logger.error(f"Error organizing file '{file.name}': {str(e)}")
                continue
                
            time.sleep(0.5)  # Rate limiting
        
        if organized_count > 0:
            logger.info(f"✓ Organized {organized_count} new file(s)")
        else:
            logger.info("No new files to organize")
            
    except Exception as e:
        logger.error(f"Error processing new files: {str(e)}")


# DRIVE ORGANIZER
class DriveOrganizer:

    def __init__(self, credentials_path: str, groq_api_key: str):
        self.drive = GoogleDriveClient(credentials_path)
        self.classifier = AIClassifier(groq_api_key)
        self.folders: Dict[str, str] = {}
        self.organized_file_ids = set()  # Track organized files

    def _should_skip_file(self, file: FileInfo) -> bool:
        """Check if file should be skipped"""
        # Skip folders
        if file.mime_type == "application/vnd.google-apps.folder":
            return True
        
        # Skip images, videos, audio
        for prefix in SKIP_MIME_PREFIXES:
            if file.mime_type.startswith(prefix):
                return True
        
        return False

    def _is_organized(self, file: FileInfo) -> bool:
        """Check if file is already in a category folder"""
        # Check if in tracking set
        if file.id in self.organized_file_ids:
            return True
        
        # Check if parent is one of our category folders
        for parent_id in file.parents:
            if parent_id in self.folders.values():
                self.organized_file_ids.add(file.id)
                return True
        
        return False

    def setup_folders(self, root_folder_id: str = "root"):
        """Create all category folders"""
        logger.info("Setting up category folders...")
        
        for category in CATEGORIES + ["Needs Review"]:
            folder_id = self.drive.create_folder(category, root_folder_id)
            if folder_id:
                self.folders[category] = folder_id
        
        logger.info(f"✓ Created/verified {len(self.folders)} category folders")

    def organize_single_file(self, file: FileInfo) -> bool:
        """Organize a single file"""
        try:
            # Download and extract content
            content_bytes = self.drive.download_file_content(file.id, file.mime_type)
            
            if content_bytes:
                content = ContentExtractor.extract(content_bytes, file.mime_type, file.name)
            else:
                content = f"Filename: {file.name}"
            
            # Classify
            classification = self.classifier.classify(file, content)
            
            # Determine destination
            destination = (
                classification.category
                if classification.confidence >= CONFIDENCE_THRESHOLD
                else "Needs Review"
            )
            
            # Move file
            destination_folder_id = self.folders.get(destination)
            if not destination_folder_id:
                logger.error(f"Destination folder not found: {destination}")
                return False
            
            success = self.drive.move_file(file.id, destination_folder_id)
            
            if success:
                self.organized_file_ids.add(file.id)
                logger.info(f"✓ Moved '{file.name}' → {destination} (confidence: {classification.confidence:.2f})")
                return True
            else:
                return False
                
        except Exception as e:
            logger.error(f"Error organizing file '{file.name}': {str(e)}")
            return False

    def organize_batch(self, root_folder_id: str = "root", dry_run: bool = False):
        """Organize all files in a folder (one-time batch operation)"""
        logger.info(f"Starting batch organization (dry_run={dry_run})")
        
        # Setup folders
        self.setup_folders(root_folder_id)
        
        # Get all files
        files = self.drive.list_files(root_folder_id)
        
        stats = {
            'total': len(files),
            'organized': 0,
            'needs_review': 0,
            'skipped': 0,
            'errors': 0
        }
        
        logger.info(f"Processing {stats['total']} files...")
        
        for idx, file in enumerate(files, 1):
            logger.info(f"[{idx}/{stats['total']}] Processing: {file.name}")
            
            # Skip if should be skipped
            if self._should_skip_file(file):
                logger.info(f"  → Skipped ({file.mime_type})")
                stats['skipped'] += 1
                continue
            
            # Skip if already organized
            if self._is_organized(file):
                logger.info(f"  → Already organized")
                stats['skipped'] += 1
                continue
            
            try:
                if not dry_run:
                    success = self.organize_single_file(file)
                    if success:
                        stats['organized'] += 1
                    else:
                        stats['errors'] += 1
                else:
                    # Dry run - just classify
                    content_bytes = self.drive.download_file_content(file.id, file.mime_type)
                    content = ContentExtractor.extract(content_bytes, file.mime_type, file.name) if content_bytes else f"Filename: {file.name}"
                    classification = self.classifier.classify(file, content)
                    
                    destination = classification.category if classification.confidence >= CONFIDENCE_THRESHOLD else "Needs Review"
                    logger.info(f"  → [DRY RUN] Would move to '{destination}' (confidence: {classification.confidence:.2f})")
                    stats['organized'] += 1
                    
            except Exception as e:
                logger.error(f"  → Error: {str(e)}")
                stats['errors'] += 1
                
            time.sleep(0.5)  # Rate limiting
        
        # Print summary
        logger.info("\n" + "="*60)
        logger.info("BATCH ORGANIZATION SUMMARY")
        logger.info("="*60)
        logger.info(f"Total files: {stats['total']}")
        logger.info(f"Organized: {stats['organized']}")
        logger.info(f"Skipped: {stats['skipped']}")
        logger.info(f"Errors: {stats['errors']}")
        logger.info("="*60 + "\n")


# WEBHOOK SERVER
class WebhookServer:
    """Manages webhook subscription and server"""
    
    def __init__(self, organizer: DriveOrganizer, webhook_url: str):
        self.organizer = organizer
        self.webhook_url = webhook_url
        self.channel_id = str(uuid.uuid4())
        self.watch_response = None
        
    def start_watching(self, folder_id: str = "root"):
        """Subscribe to Drive changes"""
        try:
            # Watch request
            body = {
                'id': self.channel_id,
                'type': 'web_hook',
                'address': self.webhook_url,
                'expiration': int((time.time() + 604800) * 1000)  # 7 days (max allowed)
            }
            
            # Start watching
            self.watch_response = self.organizer.drive.service.files().watch(
                fileId=folder_id,
                body=body,
                supportsAllDrives=True
            ).execute()
            
            expiration_time = datetime.fromtimestamp(int(self.watch_response['expiration'])/1000)
            
            logger.info("="*60)
            logger.info("✓ WEBHOOK SUBSCRIPTION ACTIVE")
            logger.info("="*60)
            logger.info(f"Channel ID: {self.channel_id}")
            logger.info(f"Resource ID: {self.watch_response['resourceId']}")
            logger.info(f"Expires: {expiration_time}")
            logger.info("="*60 + "\n")
            
            return True
            
        except Exception as e:
            logger.error(f"Error starting webhook: {str(e)}")
            return False
            
    def stop_watching(self):
        """Unsubscribe from Drive changes"""
        if self.watch_response:
            try:
                self.organizer.drive.service.channels().stop(
                    body={
                        'id': self.channel_id,
                        'resourceId': self.watch_response['resourceId']
                    }
                ).execute()
                logger.info("✓ Stopped webhook subscription")
            except Exception as e:
                logger.error(f"Error stopping webhook: {str(e)}")
    
    def run_server(self, host: str = '0.0.0.0', port: int = 5000):
        """Start Flask webhook server"""
        logger.info(f"Starting webhook server on {host}:{port}")
        logger.info(f"Webhook endpoint: {self.webhook_url}")
        logger.info("Press Ctrl+C to stop\n")
        
        try:
            app.run(host=host, port=port, debug=False)
        except KeyboardInterrupt:
            logger.info("\nShutting down...")
            self.stop_watching()


# MAIN
if __name__ == "__main__":

    CREDENTIALS_FILE = "credentials.json"
    GROQ_API_KEY = os.getenv("GROQ_API_KEY")

    if not GROQ_API_KEY:
        print("Error: GROQ_API_KEY environment variable not set")
        print("Set it using: export GROQ_API_KEY='your-api-key'")
        exit(1)

    print("\n" + "="*60)
    print("AI-POWERED GOOGLE DRIVE ORGANIZER")
    print("="*60)
    print("1. One-time batch organization (dry run)")
    print("2. One-time batch organization (live)")
    print("3. Real-time auto-organization (webhook server)")
    print("="*60)
    
    mode = input("\nChoose mode (1/2/3): ").strip()
    
    # Initialize organizer
    organizer = DriveOrganizer(CREDENTIALS_FILE, GROQ_API_KEY)
    organizer_instance = organizer  # Global reference for webhook
    
    if mode in ["1", "2"]:
        # Batch organization
        dry_run = (mode == "1")
        organizer.organize_batch(root_folder_id="root", dry_run=dry_run)
        
    elif mode == "3":
        # Real-time webhook mode
        print("\n" + "="*60)
        print("REAL-TIME WEBHOOK SETUP")
        print("="*60)
        print("Requirements:")
        print("1. Public webhook URL (use ngrok or deploy to cloud)")
        print("2. Domain verified in Google Cloud Console")
        print("3. Port 5000 available")
        print("="*60 + "\n")
        
        webhook_url = input("Enter your public webhook URL (e.g., https://your-domain.com/webhook/drive): ").strip()
        
        if not webhook_url:
            print("Error: Webhook URL required")
            exit(1)
        
        if not webhook_url.startswith("https://"):
            print("Warning: Google Drive webhooks require HTTPS")
        
        # Setup folders first
        organizer.setup_folders("root")
        
        # Start webhook subscription
        webhook_server = WebhookServer(organizer, webhook_url)
        
        if webhook_server.start_watching("root"):
            # Run Flask server
            webhook_server.run_server()
        else:
            print("Failed to start webhook subscription")
            exit(1)
    
    else:
        print("Invalid mode")